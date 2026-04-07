#!/usr/bin/env python3
"""Exporta planilha .xlsx → web/public/data/dashboard-bundle.json.

Uso:
  python3 scripts/export_bundle.py
  python3 scripts/export_bundle.py "/caminho/planilha.xlsx"

Planilhas suportadas:
  • Legado: aba «dados brutos dashboard» (ou «Página1»), 6 colunas + cabeçalho.
  • Google export: cabeçalho com id_medidor (5 colunas, sem instalador).
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    raise SystemExit("Instale openpyxl: pip install openpyxl")

from dashboard_bundle_builder import (
    OUT_DEFAULT,
    SHEET_TRIES_LOCAL,
    build_bundle,
    eventos_from_grid,
    write_bundle,
)

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "help xpezinha novo.xlsx"


def pick_worksheet(wb: openpyxl.Workbook) -> tuple[object, str]:
    for name in SHEET_TRIES_LOCAL:
        if name in wb.sheetnames:
            return wb[name], name
    first = wb.sheetnames[0]
    return wb[first], first


def main() -> None:
    xlsx = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.is_file():
        raise SystemExit(f"Planilha não encontrada: {xlsx}")

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    ws, sheet_used = pick_worksheet(wb)

    all_rows: list[list[object]] = []
    for row in ws.iter_rows(min_row=1, values_only=True):
        if row is None:
            continue
        all_rows.append(list(row))

    wb.close()

    eventos = eventos_from_grid(all_rows)
    bundle = build_bundle(eventos, xlsx.name)
    out = write_bundle(bundle)
    n_med = len(bundle["frota"]["idsMedidoresObservados"])
    print(f"OK -> {out} ({len(eventos)} eventos, {n_med} medidores) · aba «{sheet_used}» · {xlsx.name}")


if __name__ == "__main__":
    main()
